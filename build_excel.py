import csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ---------- Sheet 1: Raw transaction log ----------
ws = wb.active
ws.title = "Transactions"

header_font = Font(name="Arial", bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
normal_font = Font(name="Arial", size=10)
thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

headers = ["Date", "Category", "Type", "Amount"]
for c, h in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")

with open("raw_expenses.csv") as f:
    reader = csv.reader(f)
    next(reader)
    r = 2
    for row in reader:
        d, cat, typ, amt = row
        ws.cell(row=r, column=1, value=datetime.strptime(d, "%Y-%m-%d").date()).number_format = "yyyy-mm-dd"
        ws.cell(row=r, column=2, value=cat)
        ws.cell(row=r, column=3, value=typ)
        ws.cell(row=r, column=4, value=float(amt)).number_format = '$#,##0.00'
        for c in range(1, 5):
            ws.cell(row=r, column=c).font = normal_font
            ws.cell(row=r, column=c).border = border
        r += 1

last_row = r - 1
ws.column_dimensions["A"].width = 14
ws.column_dimensions["B"].width = 16
ws.column_dimensions["C"].width = 12
ws.column_dimensions["D"].width = 12
ws.freeze_panes = "A2"

# Total row with a real formula
total_row = r
ws.cell(row=total_row, column=3, value="Total").font = Font(name="Arial", bold=True)
tc = ws.cell(row=total_row, column=4, value=f"=SUM(D2:D{last_row})")
tc.number_format = '$#,##0.00'
tc.font = Font(name="Arial", bold=True)

# ---------- Sheet 2: Monthly Summary (SUMIFS formulas) ----------
ws2 = wb.create_sheet("Monthly Summary")

categories = ["Food", "Transport", "Rent", "Utilities", "Entertainment", "Shopping", "Health", "Education"]
months = ["2025-09","2025-10","2025-11","2025-12","2026-01","2026-02","2026-03",
          "2026-04","2026-05","2026-06","2026-07","2026-08"]

ws2.cell(row=1, column=1, value="Category").font = header_font
ws2.cell(row=1, column=1).fill = header_fill
for i, m in enumerate(months, start=2):
    cell = ws2.cell(row=1, column=i, value=m)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")

for ci, cat in enumerate(categories, start=2):
    ws2.cell(row=ci, column=1, value=cat).font = Font(name="Arial", bold=True)
    for mi, m in enumerate(months, start=2):
        year, month = m.split("-")
        col_letter = get_column_letter(mi)
        # SUMIFS: sum Amount where Category matches and Date falls within that month
        formula = (
            f'=SUMIFS(Transactions!$D$2:$D${last_row},'
            f'Transactions!$B$2:$B${last_row},$A{ci},'
            f'Transactions!$A$2:$A${last_row},">="&DATE({year},{month},1),'
            f'Transactions!$A$2:$A${last_row},"<"&EDATE(DATE({year},{month},1),1))'
        )
        cell = ws2.cell(row=ci, column=mi, value=formula)
        cell.number_format = '$#,##0.00'
        cell.font = normal_font
        cell.border = border

# Monthly total row (sum of category column)
total_r = len(categories) + 2
ws2.cell(row=total_r, column=1, value="Monthly Total").font = Font(name="Arial", bold=True)
for mi in range(2, len(months) + 2):
    col_letter = get_column_letter(mi)
    cell = ws2.cell(row=total_r, column=mi, value=f"=SUM({col_letter}2:{col_letter}{total_r-1})")
    cell.number_format = '$#,##0.00'
    cell.font = Font(name="Arial", bold=True)
    cell.border = border

ws2.column_dimensions["A"].width = 16
for i in range(2, len(months) + 2):
    ws2.column_dimensions[get_column_letter(i)].width = 12
ws2.freeze_panes = "B2"

wb.save("Expense_Tracker.xlsx")
print("Saved Expense_Tracker.xlsx")
